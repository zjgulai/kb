#!/usr/bin/env python3
"""Generate parser unit manifests for the full consultant-role source register.

Boundary: local parser coverage only. The manifest records locators and length
metadata, not source text. It does not call a provider or ingest a live KB.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import xml.etree.ElementTree as ET
from collections import Counter
from pathlib import Path
from typing import Any
from zipfile import ZipFile

import openpyxl
from docx import Document
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
REGISTER_PATH = ROOT / "drafts/analysis/consultant-role-kb-full-source-register-20260619.csv"
MANIFEST_OUT = ROOT / "tmp/consultant-role-kb-parser-unit-manifest-20260619.jsonl"
SUMMARY_OUT = ROOT / "tmp/consultant-role-kb-parser-unit-manifest-summary-20260619.json"
REPORT_OUT = ROOT / "drafts/analysis/consultant-role-kb-parser-unit-manifest-report-20260619.md"

PPT_NS = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}


def read_register(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8") as f:
        return list(csv.DictReader(f))


def compact_text(text: Any) -> str:
    return re.sub(r"\s+", " ", str(text or "").strip())


def unit(locator_type: str, locator: str, text_length: int, empty: bool, **extra: Any) -> dict[str, Any]:
    row = {
        "locator_type": locator_type,
        "locator": locator,
        "text_length": int(text_length),
        "empty": bool(empty),
    }
    row.update(extra)
    return row


def parse_pdf(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    units: list[dict[str, Any]] = []
    reader = PdfReader(str(path))
    for idx, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception as exc:  # noqa: BLE001 - parser diagnostics should continue per page.
            warnings.append(f"page:{idx}:extract_text_failed:{type(exc).__name__}")
            text = ""
        text = compact_text(text)
        units.append(unit("page", f"page:{idx}", len(text), not bool(text)))
    return units, warnings


def parse_pptx(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    units: list[dict[str, Any]] = []
    with ZipFile(path) as z:
        slide_names = sorted(
            [name for name in z.namelist() if re.match(r"ppt/slides/slide\d+\.xml$", name)],
            key=lambda name: int(re.search(r"slide(\d+)\.xml", name).group(1)),
        )
        for idx, slide_name in enumerate(slide_names, start=1):
            try:
                root = ET.fromstring(z.read(slide_name))
                texts = [node.text for node in root.findall(".//a:t", PPT_NS) if node.text]
                text = compact_text(" ".join(texts))
            except Exception as exc:  # noqa: BLE001
                warnings.append(f"slide:{idx}:parse_failed:{type(exc).__name__}")
                text = ""
            units.append(unit("slide", f"slide:{idx}", len(text), not bool(text)))
    return units, warnings


def parse_docx(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    units: list[dict[str, Any]] = []
    doc = Document(str(path))
    for idx, paragraph in enumerate(doc.paragraphs, start=1):
        text = compact_text(paragraph.text)
        if text:
            units.append(unit("paragraph", f"paragraph:{idx}", len(text), False))
    return units, warnings


def parse_xlsx(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    units: list[dict[str, Any]] = []
    workbook = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        if hasattr(worksheet, "reset_dimensions"):
            try:
                worksheet.reset_dimensions()
            except Exception as exc:  # noqa: BLE001
                warnings.append(f"sheet:{sheet_name}:reset_dimensions_failed:{type(exc).__name__}")
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = [compact_text(value) for value in row if value not in (None, "")]
            if not values:
                continue
            text_length = len(" ".join(values))
            units.append(
                unit(
                    "sheet_row",
                    f"sheet:{sheet_name}#row:{row_idx}",
                    text_length,
                    text_length == 0,
                    sheet_name=sheet_name,
                    row_index=row_idx,
                    non_empty_cell_count=len(values),
                )
            )
    return units, warnings


def parse_csv(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    units: list[dict[str, Any]] = []
    try:
        f = path.open(encoding="utf-8-sig", newline="")
        close_file = True
    except UnicodeDecodeError:
        f = path.open(encoding="latin-1", newline="")
        close_file = True
    try:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            values = [compact_text(value) for value in row if value not in (None, "")]
            if not values:
                continue
            units.append(
                unit(
                    "csv_row",
                    f"row:{row_idx}",
                    len(" ".join(values)),
                    False,
                    row_index=row_idx,
                    non_empty_cell_count=len(values),
                )
            )
    finally:
        if close_file:
            f.close()
    return units, warnings


def strip_html(text: str) -> str:
    text = re.sub(r"<script\b.*?</script>", " ", text, flags=re.IGNORECASE | re.DOTALL)
    text = re.sub(r"<style\b.*?</style>", " ", text, flags=re.IGNORECASE | re.DOTALL)
    text = re.sub(r"<[^>]+>", " ", text)
    return compact_text(text)


def parse_epub(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    units: list[dict[str, Any]] = []
    with ZipFile(path) as z:
        html_names = sorted(
            [
                name
                for name in z.namelist()
                if name.lower().endswith((".html", ".xhtml", ".htm")) and not name.endswith("/")
            ]
        )
        for idx, name in enumerate(html_names, start=1):
            try:
                raw = z.read(name).decode("utf-8", errors="ignore")
                text = strip_html(raw)
            except Exception as exc:  # noqa: BLE001
                warnings.append(f"section:{idx}:parse_failed:{type(exc).__name__}")
                text = ""
            units.append(unit("section", f"section:{idx}", len(text), not bool(text), file_name=name))
    return units, warnings


def parse_source(row: dict[str, str]) -> dict[str, Any]:
    path = Path(row["source_uri"])
    suffix = path.suffix.lower()
    warnings: list[str] = []
    parse_error = ""
    units: list[dict[str, Any]] = []

    try:
        if suffix == ".pdf":
            units, warnings = parse_pdf(path)
        elif suffix == ".pptx":
            units, warnings = parse_pptx(path)
        elif suffix == ".docx":
            units, warnings = parse_docx(path)
        elif suffix == ".xlsx":
            units, warnings = parse_xlsx(path)
        elif suffix == ".csv":
            units, warnings = parse_csv(path)
        elif suffix == ".epub":
            units, warnings = parse_epub(path)
        else:
            parse_error = f"unsupported_suffix:{suffix or 'none'}"
    except Exception as exc:  # noqa: BLE001
        parse_error = f"{type(exc).__name__}:{exc}"
        units = []

    unit_count = len(units)
    empty_units = sum(1 for item in units if item["empty"])
    return {
        "source_id": row["source_id"],
        "source_title": row["source_title"],
        "source_uri": row["source_uri"],
        "parser_route": row.get("parser_route") or row.get("source_suffix") or suffix,
        "source_suffix": suffix,
        "workspace": row["workspace"],
        "license_status": row["license_status"],
        "evidence_grade": row["evidence_grade"],
        "unit_count": unit_count,
        "empty_unit_count": empty_units,
        "empty_unit_rate": round(empty_units / unit_count, 4) if unit_count else None,
        "warnings": warnings[:50],
        "warning_count": len(warnings),
        "parse_error": parse_error,
        "units": units,
    }


def summarize(manifests: list[dict[str, Any]]) -> dict[str, Any]:
    total_sources = len(manifests)
    parse_error_count = sum(1 for item in manifests if item["parse_error"])
    total_units = sum(item["unit_count"] for item in manifests)
    empty_units = sum(item["empty_unit_count"] for item in manifests)
    by_suffix = Counter(item["source_suffix"] for item in manifests)
    by_locator = Counter(
        unit["locator_type"] for manifest in manifests for unit in manifest["units"]
    )
    return {
        "created_at": "2026-06-19",
        "scope": "full consultant-role parser unit manifest",
        "production_impact": "production unchanged",
        "provider_call_boundary": "no KB provider call",
        "implementation_status": "local parser manifests only; no live KB ingestion",
        "source_count": total_sources,
        "parse_success_count": total_sources - parse_error_count,
        "parse_error_count": parse_error_count,
        "total_unit_count": total_units,
        "empty_unit_count": empty_units,
        "empty_unit_rate": round(empty_units / total_units, 4) if total_units else None,
        "by_suffix": dict(sorted(by_suffix.items())),
        "by_locator_type": dict(sorted(by_locator.items())),
        "sources_with_warnings": sum(1 for item in manifests if item["warning_count"]),
    }


def write_outputs(manifests: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    MANIFEST_OUT.write_text(
        "\n".join(json.dumps(item, ensure_ascii=False, sort_keys=True) for item in manifests) + "\n",
        encoding="utf-8",
    )
    SUMMARY_OUT.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    report = f"""---
title: "Consultant Role KB Parser Unit Manifest Report"
status: "draft"
created_at: "2026-06-19"
source_documents:
  - "drafts/analysis/consultant-role-kb-full-source-register-20260619.csv"
scope: "parser unit manifest coverage for full consultant-agent source register"
production_impact: "production unchanged"
provider_call_boundary: "no KB provider call"
implementation_status: "local parser manifests only; no live KB ingestion"
---

# Consultant Role KB Parser Unit Manifest Report

## 0. Boundary

The manifest records structural locators, unit counts, text lengths, and parser
diagnostics. It intentionally does not persist raw source text.

## 1. Outputs

| artifact | path |
|---|---|
| unit manifest JSONL | `tmp/consultant-role-kb-parser-unit-manifest-20260619.jsonl` |
| summary JSON | `tmp/consultant-role-kb-parser-unit-manifest-summary-20260619.json` |

## 2. Coverage Metrics

| metric | value |
|---|---:|
| source_count | {summary['source_count']} |
| parse_success_count | {summary['parse_success_count']} |
| parse_error_count | {summary['parse_error_count']} |
| total_unit_count | {summary['total_unit_count']} |
| empty_unit_count | {summary['empty_unit_count']} |
| empty_unit_rate | {summary['empty_unit_rate']} |
| sources_with_warnings | {summary['sources_with_warnings']} |
| provider_call_count | 0 |
| live_kb_write_count | 0 |

## 3. Units By Locator Type

| locator_type | count |
|---|---:|
"""
    for locator_type, count in summary["by_locator_type"].items():
        report += f"| {locator_type} | {count} |\n"

    report += """
## 4. Sources By Suffix

| suffix | count |
|---|---:|
"""
    for suffix, count in summary["by_suffix"].items():
        report += f"| {suffix or 'unknown'} | {count} |\n"

    failed = [item for item in manifests if item["parse_error"]]
    report += """
## 5. Parse Errors

"""
    if failed:
        for item in failed:
            report += f"- `{item['source_id']}` {item['parse_error']}\n"
    else:
        report += "- None.\n"

    report += """
## 6. Interpretation

Fact: the full registered source set now has structural parser manifests that
can be used by card QA and later extraction batches.

Inference: source registration and parser coverage are ready for batch-level
full extraction planning, subject to legal/source-owner gates.

Unknown: this manifest does not prove extracted card quality or answer quality.
"""
    REPORT_OUT.write_text(report, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--register", default=str(REGISTER_PATH))
    parser.add_argument("--limit", type=int, default=0)
    args = parser.parse_args()

    rows = read_register(Path(args.register))
    if args.limit:
        rows = rows[: args.limit]
    manifests = [parse_source(row) for row in rows]
    summary = summarize(manifests)
    write_outputs(manifests, summary)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
