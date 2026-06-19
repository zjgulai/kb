#!/usr/bin/env python3
"""Add source anchors to consultant role KB sample cards and score anchor precision.

Boundary: local PoC only, no provider call, no live KB ingestion,
production unchanged.
"""

from __future__ import annotations

import csv
import json
import re
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any
from zipfile import ZipFile

import openpyxl
from docx import Document
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
CARD_PATH = ROOT / "tmp/consultant-role-kb-card-samples-20260619.jsonl"
SOURCE_REGISTER_PATH = ROOT / "drafts/analysis/consultant-role-kb-source-register-candidates-20260619.csv"

ANCHORED_CARD_OUT = ROOT / "tmp/consultant-role-kb-card-samples-anchored-20260619.jsonl"
ANCHOR_EVAL_OUT = ROOT / "tmp/consultant-role-kb-citation-anchor-eval-20260619.json"
REPORT_OUT = ROOT / "drafts/analysis/consultant-role-kb-citation-anchor-poc-report-20260619.md"

PPT_NS = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
STOPWORDS = {
    "a",
    "an",
    "and",
    "are",
    "as",
    "at",
    "be",
    "by",
    "for",
    "from",
    "in",
    "into",
    "is",
    "it",
    "of",
    "on",
    "or",
    "our",
    "source",
    "the",
    "this",
    "to",
    "use",
    "with",
}


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def read_sources(path: Path) -> dict[str, dict[str, str]]:
    with path.open(encoding="utf-8") as f:
        return {row["source_id"]: row for row in csv.DictReader(f)}


def normalize(text: Any) -> str:
    return re.sub(r"\s+", " ", str(text or "").strip().lower())


def safe_label(text: str, max_len: int = 96) -> str:
    text = re.sub(r"\s+", " ", text.strip())
    return text[:max_len]


def text_terms(text: str) -> set[str]:
    english = re.findall(r"[a-zA-Z0-9][a-zA-Z0-9&/-]*", text.lower())
    chinese = re.findall(r"[\u4e00-\u9fff]{2,}", text)
    return {term for term in english + chinese if len(term) >= 2 and term not in STOPWORDS}


def flatten(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, (int, float, bool)):
        return str(value)
    if isinstance(value, list):
        return " ".join(flatten(item) for item in value)
    if isinstance(value, dict):
        return " ".join(flatten(val) for val in value.values())
    return str(value)


def card_queries(card: dict[str, Any], source: dict[str, str]) -> list[str]:
    fields = [
        card.get("method_name"),
        card.get("diagnostic_name"),
        card.get("dimension_names"),
        card.get("deliverable_type"),
        card.get("purpose"),
        card.get("sections"),
        card.get("outputs"),
        card.get("kpi_name"),
        card.get("function_or_industry"),
        card.get("group"),
        card.get("term"),
        card.get("expansion_or_mapping"),
        card.get("sheet_name"),
        card.get("source_structure_sample"),
    ]
    raw_terms: list[str] = []
    for field in fields:
        if isinstance(field, list):
            raw_terms.extend(str(item) for item in field if str(item).strip())
        else:
            text = str(field or "").strip()
            if text:
                raw_terms.append(text)

    # Put the most card-specific terms first.
    priority = [
        card.get("kpi_name"),
        card.get("term"),
        card.get("expansion_or_mapping"),
        card.get("method_name"),
        card.get("diagnostic_name"),
        card.get("deliverable_type"),
    ]
    ordered: list[str] = []
    seen: set[str] = set()
    for text in [*priority, *raw_terms]:
        norm = normalize(text)
        if norm and norm not in seen and len(norm) >= 2:
            ordered.append(str(text))
            seen.add(norm)
    return ordered[:24]


def match_score(queries: list[str], unit_text: str) -> tuple[float, list[str]]:
    unit_norm = normalize(unit_text)
    matched: list[str] = []
    exact_score = 0.0
    for idx, query in enumerate(queries):
        q_norm = normalize(query)
        if len(q_norm) >= 4 and q_norm in unit_norm:
            matched.append(safe_label(query, 80))
            exact_score += 2.0 if idx < 3 else 1.0

    query_terms = set()
    for query in queries:
        query_terms.update(text_terms(query))
    unit_terms = text_terms(unit_norm)
    overlap = sorted(query_terms & unit_terms)
    for term in overlap[:8]:
        if term not in matched:
            matched.append(term)

    overlap_score = min(len(overlap), 10) / 10
    score = exact_score + overlap_score
    return score, matched[:10]


def parse_pptx(path: Path) -> list[dict[str, Any]]:
    units = []
    with ZipFile(path) as z:
        slide_names = sorted(
            [name for name in z.namelist() if re.match(r"ppt/slides/slide\d+\.xml$", name)],
            key=lambda name: int(re.search(r"slide(\d+)\.xml", name).group(1)),
        )
        for idx, slide_name in enumerate(slide_names, start=1):
            root = ET.fromstring(z.read(slide_name))
            texts = [node.text for node in root.findall(".//a:t", PPT_NS) if node.text]
            text = " ".join(texts)
            title = next((safe_label(item) for item in texts if item and item.strip()), f"slide {idx}")
            units.append({"locator_type": "slide", "locator": f"slide:{idx}", "label": title, "text": text})
    return units


def parse_pdf(path: Path) -> list[dict[str, Any]]:
    units = []
    reader = PdfReader(str(path))
    for idx, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        label = safe_label(text, 96) or f"page {idx}"
        units.append({"locator_type": "page", "locator": f"page:{idx}", "label": label, "text": text})
    return units


def parse_docx(path: Path) -> list[dict[str, Any]]:
    units = []
    doc = Document(str(path))
    for idx, paragraph in enumerate(doc.paragraphs, start=1):
        text = paragraph.text.strip()
        if text:
            units.append(
                {
                    "locator_type": "paragraph",
                    "locator": f"paragraph:{idx}",
                    "label": safe_label(text, 96),
                    "text": text,
                }
            )
    return units


def parse_xlsx(path: Path, preferred_sheet: str | None = None) -> list[dict[str, Any]]:
    units = []
    workbook = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    if preferred_sheet and preferred_sheet in sheet_names:
        sheet_names = [preferred_sheet]
    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = [str(value).strip() for value in row if value not in (None, "")]
            if not values:
                continue
            text = " ".join(values)
            units.append(
                {
                    "locator_type": "sheet_row",
                    "locator": f"sheet:{sheet_name}#row:{row_idx}",
                    "label": safe_label(text, 96),
                    "text": text,
                    "sheet_name": sheet_name,
                    "row_index": row_idx,
                }
            )
    return units


def parse_csv(path: Path) -> list[dict[str, Any]]:
    units = []
    try:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle))
    except UnicodeDecodeError:
        with path.open(encoding="latin-1", newline="") as handle:
            rows = list(csv.reader(handle))
    for row_idx, row in enumerate(rows, start=1):
        values = [str(value).strip() for value in row if value not in (None, "")]
        if not values:
            continue
        text = " ".join(values)
        units.append(
            {
                "locator_type": "csv_row",
                "locator": f"row:{row_idx}",
                "label": safe_label(text, 96),
                "text": text,
                "row_index": row_idx,
            }
        )
    return units


def load_units(path: Path, card: dict[str, Any]) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".pptx":
        return parse_pptx(path)
    if suffix == ".pdf":
        return parse_pdf(path)
    if suffix == ".docx":
        return parse_docx(path)
    if suffix == ".xlsx":
        return parse_xlsx(path, card.get("sheet_name"))
    if suffix == ".csv":
        return parse_csv(path)
    return []


def select_anchors(card: dict[str, Any], source: dict[str, str], units: list[dict[str, Any]]) -> list[dict[str, Any]]:
    queries = card_queries(card, source)
    scored = []
    for unit in units:
        score, matched_terms = match_score(queries, unit["text"])
        if score <= 0:
            continue
        scored.append((score, unit, matched_terms))
    scored.sort(key=lambda item: item[0], reverse=True)

    anchors = []
    for score, unit, matched_terms in scored[:5]:
        confidence = "high" if score >= 2.0 else "medium" if score >= 1.0 else "low"
        anchors.append(
            {
                "source_id": card["source_id"],
                "locator_type": unit["locator_type"],
                "locator": unit["locator"],
                "anchor_label": unit["label"],
                "matched_terms": matched_terms,
                "anchor_match_score": round(score, 4),
                "anchor_confidence": confidence,
            }
        )
    return anchors


def score_anchor(card: dict[str, Any], anchors: list[dict[str, Any]]) -> dict[str, Any]:
    if not anchors:
        return {
            "card_id": card["card_id"],
            "source_id": card["source_id"],
            "source_modality": Path(card["source_uri"]).suffix.lower().lstrip("."),
            "anchor_status": "missing",
            "citation_precision_ready": False,
            "best_anchor": None,
            "anchor_count": 0,
        }
    best = anchors[0]
    ready = best["anchor_confidence"] in {"high", "medium"} and best["locator_type"] in {
        "page",
        "slide",
        "sheet_row",
        "paragraph",
        "csv_row",
    }
    return {
        "card_id": card["card_id"],
        "source_id": card["source_id"],
        "source_modality": Path(card["source_uri"]).suffix.lower().lstrip("."),
        "anchor_status": "resolved" if ready else "weak",
        "citation_precision_ready": ready,
        "best_anchor": best,
        "anchor_count": len(anchors),
    }


def summarize(rows: list[dict[str, Any]]) -> dict[str, Any]:
    total = len(rows)
    ready = sum(row["citation_precision_ready"] for row in rows)
    by_modality: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_modality[row["source_modality"]].append(row)
    return {
        "card_count": total,
        "citation_precision_ready_count": ready,
        "citation_precision_ready_rate": round(ready / total, 4) if total else None,
        "anchor_status_counts": dict(Counter(row["anchor_status"] for row in rows)),
        "by_modality": {
            modality: {
                "card_count": len(items),
                "citation_precision_ready_count": sum(item["citation_precision_ready"] for item in items),
                "citation_precision_ready_rate": round(
                    sum(item["citation_precision_ready"] for item in items) / len(items), 4
                ),
                "anchor_status_counts": dict(Counter(item["anchor_status"] for item in items)),
            }
            for modality, items in sorted(by_modality.items())
        },
    }


def main() -> None:
    cards = read_jsonl(CARD_PATH)
    sources = read_sources(SOURCE_REGISTER_PATH)
    anchored_cards = []
    eval_rows = []
    unit_cache: dict[tuple[str, str], list[dict[str, Any]]] = {}

    for card in cards:
        source = sources[card["source_id"]]
        path = Path(card["source_uri"])
        preferred_sheet = card.get("sheet_name", "")
        cache_key = (str(path), preferred_sheet)
        if cache_key not in unit_cache:
            unit_cache[cache_key] = load_units(path, card)
        anchors = select_anchors(card, source, unit_cache[cache_key])
        anchored = dict(card)
        anchored["source_anchors"] = anchors
        anchored["citation_anchor_policy"] = {
            "allowed_citation_granularity": ["page", "slide", "sheet_row", "paragraph"],
            "blocked": ["long_source_text_reproduction", "source_only_citation_for_final_answer"],
            "boundary": "local PoC only; no KB provider call; production unchanged",
        }
        anchored_cards.append(anchored)
        eval_rows.append(score_anchor(card, anchors))

    ANCHORED_CARD_OUT.write_text(
        "\n".join(json.dumps(card, ensure_ascii=False, sort_keys=True) for card in anchored_cards) + "\n",
        encoding="utf-8",
    )

    summary = summarize(eval_rows)
    payload = {
        "metadata": {
            "created_at": "2026-06-19",
            "scope": "citation anchor precision PoC for consultant role KB typed cards",
            "provider_call_boundary": "no KB provider call",
            "production_impact": "production unchanged",
            "implementation_status": "local PoC only; no live KB ingestion",
            "input_cards": str(CARD_PATH.relative_to(ROOT)),
            "anchored_cards": str(ANCHORED_CARD_OUT.relative_to(ROOT)),
        },
        "summary": summary,
        "results": eval_rows,
    }
    ANCHOR_EVAL_OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    report = f"""---
title: "Consultant Role KB Citation Anchor PoC Report"
status: "draft"
created_at: "2026-06-19"
source_documents:
  - "tmp/consultant-role-kb-card-samples-20260619.jsonl"
  - "drafts/analysis/consultant-role-kb-source-register-candidates-20260619.csv"
scope: "local citation anchor precision PoC for consultant role KB typed cards"
production_impact: "production unchanged"
provider_call_boundary: "no KB provider call"
implementation_status: "local PoC only; no live KB ingestion"
---

# Consultant Role KB Citation Anchor PoC Report

## 0. Boundary

This PoC adds locator anchors to existing typed-card samples. It does not ingest into a live KB, does not call a provider, does not expand extraction scope, and does not reproduce long source text.

## 1. Anchor Schema

Each card can carry `source_anchors` with:

- `source_id`
- `locator_type`: `page`, `slide`, `sheet_row`, or `paragraph`
- `locator`: for example `page:3`, `slide:5`, `sheet:Library of KPIs by Function#row:12`
- `anchor_label`: short label only, not a long source excerpt
- `matched_terms`
- `anchor_confidence`

## 2. Outputs

| artifact | path |
|---|---|
| anchored card samples | `tmp/consultant-role-kb-card-samples-anchored-20260619.jsonl` |
| anchor eval result | `tmp/consultant-role-kb-citation-anchor-eval-20260619.json` |

## 3. Citation Precision Readiness

| metric | value |
|---|---:|
| card_count | {summary["card_count"]} |
| citation_precision_ready_count | {summary["citation_precision_ready_count"]} |
| citation_precision_ready_rate | {summary["citation_precision_ready_rate"]} |

## 4. By Modality

| modality | card_count | ready_count | ready_rate | status_counts |
|---|---:|---:|---:|---|
"""
    for modality, row in summary["by_modality"].items():
        report += (
            f"| {modality} | {row['card_count']} | {row['citation_precision_ready_count']} | "
            f"{row['citation_precision_ready_rate']} | `{json.dumps(row['anchor_status_counts'], ensure_ascii=False)}` |\n"
        )

    report += """
## 5. Interpretation

This proves the card schema can carry citation anchors at unit granularity for the current 33-card local sample.

It still does not prove answer quality or production readiness. The next retrieval/eval run should use anchored cards and score whether retrieved citations include the expected locator, not only the expected source.

## 6. Next Gate

Run anchored retrieval/citation eval using the anchored card file. Keep `source_only` citations out of final answer scoring, because source-level citation is too coarse for this PRD gate.
"""
    REPORT_OUT.write_text(report, encoding="utf-8")

    print(json.dumps({"summary": summary, "outputs": [str(ANCHORED_CARD_OUT), str(ANCHOR_EVAL_OUT), str(REPORT_OUT)]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
