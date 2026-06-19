#!/usr/bin/env python3
"""Readonly Excel table profiler for KB P1 source intake."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import warnings
from collections import Counter
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    category=UserWarning,
)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def value_kind(value: Any) -> str:
    if value is None:
        return "blank"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, (datetime, date)):
        return "date"
    return "text"


def compact_header(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())[:120]


def profile_sheet(ws: Any, max_rows: int) -> dict[str, Any]:
    reported_dimension = ws.calculate_dimension()
    reported_max_row = ws.max_row
    reported_max_col = ws.max_column

    reset_dimensions_applied = False
    try:
        ws.reset_dimensions()
        reset_dimensions_applied = True
    except Exception:
        reset_dimensions_applied = False

    nonempty_rows = 0
    max_seen_col = 0
    first_nonempty_row_index = None
    first_headers: list[str] = []
    second_headers: list[str] = []
    type_counts: Counter[str] = Counter()
    rows_scanned = 0

    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        rows_scanned = row_index
        values = list(row)
        if any(value is not None for value in values):
            nonempty_rows += 1
            max_seen_col = max(max_seen_col, len(values))
            type_counts.update(value_kind(value) for value in values)
            if first_nonempty_row_index is None:
                first_nonempty_row_index = row_index
                first_headers = [compact_header(value) for value in values[:80]]
            elif not second_headers:
                second_headers = [compact_header(value) for value in values[:80]]
        if row_index >= max_rows:
            break

    visible_headers = [header for header in first_headers if header]
    duplicate_headers = sorted(
        header for header, count in Counter(visible_headers).items() if count > 1
    )
    dimension_warning = (
        reported_dimension in {"A1:A1", "A1"}
        and (nonempty_rows > 1 or max_seen_col > 1)
    )

    return {
        "sheet_name": ws.title,
        "reported_dimension": reported_dimension,
        "reported_max_row": reported_max_row,
        "reported_max_col": reported_max_col,
        "reset_dimensions_applied": reset_dimensions_applied,
        "rows_scanned_limit": max_rows,
        "rows_scanned_actual": rows_scanned,
        "nonempty_rows_observed": nonempty_rows,
        "max_seen_col_observed": max_seen_col,
        "first_nonempty_row_index": first_nonempty_row_index,
        "first_header_cells": first_headers,
        "second_header_cells_present": bool(second_headers),
        "blank_header_cells_in_first_row": sum(1 for header in first_headers if not header),
        "duplicate_headers_in_first_row": duplicate_headers,
        "value_type_counts": dict(sorted(type_counts.items())),
        "dimension_warning": dimension_warning,
    }


def profile_workbook(path: Path, max_rows: int, max_sheets: int) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = [
            profile_sheet(sheet, max_rows)
            for sheet in workbook.worksheets[:max_sheets]
        ]
    finally:
        workbook.close()

    return {
        "path": str(path),
        "file_name": path.name,
        "size_bytes": path.stat().st_size,
        "sha256": sha256_file(path),
        "sheet_count_profiled": len(sheets),
        "sheets": sheets,
        "warnings": [
            f"dimension_warning:{sheet['sheet_name']}"
            for sheet in sheets
            if sheet["dimension_warning"]
        ],
    }


def paths_from_register(register_path: Path) -> list[Path]:
    rows = csv.DictReader(register_path.open(encoding="utf-8"))
    paths: list[Path] = []
    for row in rows:
        uri = row.get("source_uri", "")
        if uri.endswith(".xlsx"):
            paths.append(Path(uri))
    return paths


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Profile Excel workbooks for KB P1 source intake without emitting data rows."
    )
    parser.add_argument("paths", nargs="*", help="Excel workbook paths.")
    parser.add_argument("--source-register", help="CSV source register to read .xlsx paths from.")
    parser.add_argument("--output-json", help="Write profile JSON to this path.")
    parser.add_argument("--max-rows", type=int, default=20000)
    parser.add_argument("--max-sheets", type=int, default=3)
    args = parser.parse_args()

    workbook_paths = [Path(path) for path in args.paths]
    if args.source_register:
        workbook_paths.extend(paths_from_register(Path(args.source_register)))

    seen: set[Path] = set()
    unique_paths: list[Path] = []
    for path in workbook_paths:
        resolved = path.expanduser()
        if resolved not in seen:
            seen.add(resolved)
            unique_paths.append(resolved)

    missing = [str(path) for path in unique_paths if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing workbook(s): {missing}")

    result = {
        "profile_version": "kb-p1-excel-table-profile-v1",
        "generated_at": datetime.now(UTC).replace(microsecond=0).isoformat(),
        "read_policy": "readonly; headers and structural metadata only; no business data rows emitted",
        "max_rows_per_sheet": args.max_rows,
        "max_sheets_per_workbook": args.max_sheets,
        "workbook_count": len(unique_paths),
        "workbooks": [
            profile_workbook(path, args.max_rows, args.max_sheets)
            for path in unique_paths
        ],
    }

    output = json.dumps(result, ensure_ascii=False, indent=2)
    if args.output_json:
        Path(args.output_json).write_text(output + "\n", encoding="utf-8")
    else:
        print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
